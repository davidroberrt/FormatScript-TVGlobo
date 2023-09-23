""" O software "FORMAT SCRIPT" foi desenvolvido para as emissoras afiliadas da Rede Globo, 
    especificamente para o Setor de Programação da TV Paraíba e TV Cabo Branco. 
    
    Sua função é formatar as planilhas dos roteiros da Globo, 
    padronizando e diferenciando com cores os materiais da rede e locais, 
    além de realizar ajustes nos roteiros, como remover espaços em branco.
    
    Desenvolvedor: David Robert
    Software: Versão 1.2
    Tecnologias: Python 3.x | GUI Tkinter 
    Github: davidroberrt
    LinkedIn: davidroberrt
    Email: davidrobert.info@gmail.com
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox, Radiobutton
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
selected_color = "bdd6ee"  # cor padrão

# Função para selecionar a cor desejada
def select_color(color):
    global selected_color
    selected_color = color

# Função para processar o arquivo Excel
def process_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if file_path:
        try:
            # Adicionar uma mensagem de carregando
            loading_label = tk.Label(root, text="Preparando tudo para você, aguarde um momento...", font=("Helvetica", 14))
            loading_label.place(x=50, y=360)
            root.update()

            nova_altura_padrao = 3
            nova_altura_programa = 10

            workbook = openpyxl.load_workbook(file_path)

            sheet = workbook.active

            textos_colorir = ["PROGRAMA ATÉ 1 INTERVALO", "PROGRAMA ATÉ 2 INTERVALO", "PROGRAMA ATÉ 3 INTERVALO", "PROGRAMA ATÉ 4 INTERVALO", "PROGRAMA ATÉ 5 INTERVALO", "PROGRAMA ATÉ 6 INTERVALO"]

            for row in sheet.iter_rows():
                vazias = all(cell.value is None for cell in row)
                if vazias:
                    for cell in row:
                        sheet.row_dimensions[cell.row].height = nova_altura_padrao
                else:
                    for cell in row:
                        if cell.value:
                            cell_value = str(cell.value)
                            if "PROGRAMA:" in cell_value:
                                sheet.row_dimensions[cell.row].height = nova_altura_programa
                                cell.fill = PatternFill(start_color=selected_color, end_color=selected_color, fill_type="solid")

            # Ajuste de largura das colunas
            sheet.column_dimensions['AI'].width = 2.86
            sheet.column_dimensions['K'].width = 3.14
            sheet.column_dimensions['AH'].width = 5.29
            sheet.column_dimensions['I'].width = 5.57
            sheet.column_dimensions['M'].width = 1.86
            sheet.column_dimensions['O'].width = 2
            sheet.column_dimensions['V'].width = 5.86
            sheet.column_dimensions['Z'].width = 46.57
            sheet.column_dimensions['AT'].width = 30.43
            sheet.column_dimensions['BB'].width = 99.71


            # Ajuste de fontes e largura da planilha
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = Font(name="Calibri", size=14, bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal="left")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        if "GCR" in cell_value:
                            for c in row:
                                c.font = Font(size=14, bold=True, color="00b050")
                                cell.alignment = openpyxl.styles.Alignment(horizontal="left")
                            break

                        elif any(texto in cell_value for texto in textos_colorir):
                            for c in row:
                                c.fill = PatternFill(start_color=selected_color, end_color=selected_color, fill_type="solid")
                            break
            # Remover a mensagem de carregando
            loading_label.destroy()

            # Salvar o arquivo na pasta de downloads
            output_filename = 'FormatScript_' + os.path.basename(file_path)
            downloads_folder = os.path.expanduser("~/Downloads")
            output_path = os.path.join(downloads_folder, output_filename)
            workbook.save(output_path)  # Salvar o arquivo resultante
            # Adicionar uma mensagem de carregando
            salvo_label = tk.Label(root, text="Formatação concluída com sucesso! Salvo em Downloads", font=("Helvetica", 14))
            salvo_label.place(x=50, y=360)
            # Exibir mensagem de conclusão
            messagebox.showinfo("Concluído", f"O arquivo foi editado com sucesso!\nSalvo em: {output_path}")
            salvo_label.destroy()

        except Exception as e:
            # Remover a mensagem de carregando em caso de erro
            loading_label.destroy()
            messagebox.showerror("Erro", str(e))

# Cria a janela principal
root = tk.Tk()
root.title("FORMAT SCRIPT | Desenvolvido por David Robert")
root.geometry("1152x648")  # Define o tamanho da janela
root.iconbitmap("FS.ico")
root.resizable(False, False)  # Desabilita o redimensionamento da janela

# Carrega e exibe uma imagem de fundo
bg_image = Image.open("background_light.png")
bg_photo = ImageTk.PhotoImage(bg_image)
bg_label = tk.Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# Cria os botões de rádio para seleção de cor
color_var = tk.StringVar(value="bdd6ee")
color_radio1 = Radiobutton(root, text="AZUL", variable=color_var, value="bdd6ee", command=lambda: select_color("bdd6ee"), bg="#8eb9dd")
color_radio2 = Radiobutton(root, text="LARANJA", variable=color_var, value="f2dfb3", command=lambda: select_color("f2dfb3"), bg="#f2dfb3")
color_radio1.place(x=55, y=320)
color_radio2.place(x=115, y=320)

# Arredondar botão Processar e deixar ele azul
process_button = tk.Button(root, text="Carregar Arquivo", command=process_excel_file, bg="#5b5b5b", fg="white", width=30)
process_button.place(x=195, y=320)

# Menu bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Menu de configurações para modo noturno e modo claro
def change_background_light():
    bg_image_light = Image.open("background_light.png")
    bg_photo_light = ImageTk.PhotoImage(bg_image_light)
    bg_label.configure(image=bg_photo_light)
    bg_label.image = bg_photo_light

def change_background_dark():
    bg_image_dark = Image.open("background_dark.png")
    bg_photo_dark = ImageTk.PhotoImage(bg_image_dark)
    bg_label.configure(image=bg_photo_dark)
    bg_label.image = bg_photo_dark

visual_menu = tk.Menu(menu_bar, tearoff=0)
visual_menu.add_command(label="Modo Claro", command=change_background_light)
visual_menu.add_command(label="Modo Noturno", command=change_background_dark)
menu_bar.add_cascade(label="Configurações", menu=visual_menu)


# Menu de créditos
def show_credits():
    credits_text = """
    O software "FORMAT SCRIPT" é uma ferramenta desenvolvida por David Robert, para as emissoras afiliadas da Rede Globo, especificamente para o Setor de Programação da TV Paraíba e TV Cabo Branco. 
    
    Sua função é formatar as planilhas dos roteiros da Globo, padronizando e diferenciando com cores os materiais da rede e locais, além de realizar ajustes nos roteiros, como remover espaços em branco.
    
    Desenvolvedor: David Robert
    Software: Versão 1.1 Gratuito
    Tecnologias: Python 3.+ | GUI Tkinter 
    Github: davidroberrt
    LinkedIn: davidroberrt
    Email: davidrobert.info@gmail.com
    Contato: +55 (83) 99338-4466
    """
    messagebox.showinfo("Sobre", credits_text)

# Menu de créditos
credits_menu = tk.Menu(menu_bar, tearoff=0)
credits_menu.add_command(label="Informações", command=show_credits)
menu_bar.add_cascade(label="Sobre", menu=credits_menu)


# Menu de ajuda
def show_help():
    help_text = """
    Instruções de uso:

    Passo 1: Selecione uma cor para sua emissora, usando os botões de seleção de cor para destacar partes do roteiro
    Passo 2: Clique em "Carregar Arquivo", selecione o arquivo de roteiro a ser formatado e aguarde até carregar
    Passo 3: Após finalizar, o arquivo editado estará salvo na pasta "Downloads".
    """
    messagebox.showinfo("Ajuda", help_text)

help_menu = tk.Menu(menu_bar, tearoff=0)
help_menu.add_command(label="Como Usar", command=show_help)
menu_bar.add_cascade(label="Ajuda", menu=help_menu)
# Inicia o loop principal
root.mainloop()
